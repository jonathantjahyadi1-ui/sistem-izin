import os
from datetime import datetime, date, timedelta
from io import BytesIO

from flask import Blueprint, render_template, request, redirect, session, flash, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from extensions import db
from models import User
from meeting_room.models import MeetingRoom, RoomBooking


meeting_room_bp = Blueprint('meeting_room', __name__)


class BookingValidationError(Exception):
    pass


def get_current_user():
    if 'user_id' not in session:
        return None
    return User.query.get(session['user_id'])


def require_login():
    return 'user_id' in session


def is_supervisor(user):
    return bool(user and user.role in ['admin', 'hrd', 'direktur'])


def is_hrd(user):
    return bool(
        user and
        user.role in ['admin', 'hrd']
    )


def display_name(user):
    if not user:
        return '-'
    return getattr(user, 'nama_lengkap', None) or getattr(user, 'username', '-')


def get_user_by_id(user_id):
    if not user_id:
        return None
    return User.query.get(user_id)


def get_min_booking_days():
    raw_value = os.getenv('MIN_BOOKING_DAYS', '7')
    try:
        return int(raw_value)
    except ValueError:
        return 14


def parse_date(value):
    return datetime.strptime(value, '%Y-%m-%d').date()


def parse_time(value):
    return datetime.strptime(value, '%H:%M').time()


def has_schedule_conflict(room_id, meeting_date, start_time, end_time, exclude_booking_id=None, statuses=None):
    if statuses is None:
        statuses = [RoomBooking.STATUS_PENDING, RoomBooking.STATUS_APPROVED]

    query = RoomBooking.query.filter(
        RoomBooking.room_id == room_id,
        RoomBooking.meeting_date == meeting_date,
        RoomBooking.status.in_(statuses),
    )

    if exclude_booking_id:
        query = query.filter(RoomBooking.id != exclude_booking_id)

    for existing in query.all():
        if existing.start_time < end_time and existing.end_time > start_time:
            return True

    return False


def validate_booking(room_id, meeting_date, start_time, end_time, participant_count, exclude_booking_id=None):
    min_days = get_min_booking_days()

    if (meeting_date - date.today()).days < min_days:
        raise BookingValidationError(f'Booking harus diajukan minimal H-{min_days} sebelum tanggal meeting.')

    if end_time <= start_time:
        raise BookingValidationError('Jam selesai harus lebih besar dari jam mulai.')

    room = MeetingRoom.query.get(room_id)
    if not room:
        raise BookingValidationError('Ruangan tidak ditemukan.')

    if not room.is_active:
        raise BookingValidationError('Ruangan sedang tidak aktif.')

    if participant_count < 1:
        raise BookingValidationError('Jumlah peserta minimal 1 orang.')

    if participant_count > room.capacity:
        raise BookingValidationError(f'Jumlah peserta melebihi kapasitas ruangan. Kapasitas maksimal {room.capacity} orang.')

    if has_schedule_conflict(room_id, meeting_date, start_time, end_time, exclude_booking_id=exclude_booking_id):
        raise BookingValidationError('Ruangan sudah dibooking pada tanggal dan jam tersebut.')


@meeting_room_bp.context_processor
def inject_meeting_room_helpers():
    return {
        'meeting_get_user': get_user_by_id,
        'meeting_display_name': display_name,
        'meeting_is_supervisor': is_supervisor,
        'meeting_is_hrd': is_hrd,
    }

@meeting_room_bp.route('/')
def meeting_room_home():
    if not require_login():
        return redirect('/login')

    user = get_current_user()

    if not user:
        session.clear()
        return redirect('/login')

    return redirect('/meeting-room/list')

@meeting_room_bp.route('/list')
def list_booking():
    if not require_login():
        return redirect('/login')

    user = get_current_user()

    if not user:
        session.clear()
        return redirect('/login')

    status_filter = request.args.get('status', '').strip()
    tanggal_filter = request.args.get('tanggal', '').strip()
    room_id_filter = request.args.get('room_id', type=int)

    query = RoomBooking.query.order_by(
        RoomBooking.created_at.desc()
    )

    # Karyawan hanya melihat booking sendiri.
    # Admin / HRD / Direktur bisa melihat semua.
    if not is_supervisor(user):
        query = query.filter(
            RoomBooking.user_id == user.id
        )

    if status_filter:
        query = query.filter(
            RoomBooking.status == status_filter
        )

    if tanggal_filter:
        try:
            meeting_date = datetime.strptime(
                tanggal_filter,
                '%Y-%m-%d'
            ).date()

            query = query.filter(
                RoomBooking.meeting_date == meeting_date
            )

        except ValueError:
            flash(
                'Format tanggal filter tidak valid.',
                'warning'
            )

    if room_id_filter:
        query = query.filter(
            RoomBooking.room_id == room_id_filter
        )

    bookings = query.all()

    rooms = MeetingRoom.query.filter_by(
        is_active=True
    ).order_by(
        MeetingRoom.room_name.asc()
    ).all()

    return render_template(
        'meeting_room/list.html',
        user=user,
        bookings=bookings,
        rooms=rooms,
        statuses=RoomBooking.VALID_STATUSES,
        status_filter=status_filter,
        tanggal_filter=tanggal_filter,
        room_id_filter=room_id_filter,
    )


@meeting_room_bp.route('/form', methods=['GET', 'POST'])
def create_booking():
    if not require_login():
        return redirect('/login')

    user = get_current_user()
    rooms = MeetingRoom.get_active_rooms()
    min_days = get_min_booking_days()

    if request.method == 'POST':
        try:
            title = request.form.get('title', '').strip()
            room_id = request.form.get('room_id', type=int)
            meeting_date = parse_date(request.form.get('meeting_date', ''))
            start_time = parse_time(request.form.get('start_time', ''))
            end_time = parse_time(request.form.get('end_time', ''))
            participant_count = request.form.get('participant_count', type=int)
            purpose = request.form.get('purpose', '').strip()
            notes = request.form.get('notes', '').strip()
            division = getattr(user, 'divisi', None) or request.form.get('division', '').strip() or '-'

            if not all([title, room_id, meeting_date, start_time, end_time, participant_count, purpose]):
                flash('Semua field wajib harus diisi.', 'danger')
                return render_template('meeting_room/form.html', user=user, rooms=rooms, min_days=min_days)

            validate_booking(room_id, meeting_date, start_time, end_time, participant_count)

            booking = RoomBooking(
                user_id=user.id,
                room_id=room_id,
                title=title,
                division=division,
                meeting_date=meeting_date,
                start_time=start_time,
                end_time=end_time,
                participant_count=participant_count,
                purpose=purpose,
                notes=notes,
                status=RoomBooking.STATUS_PENDING,
            )

            db.session.add(booking)
            db.session.commit()

            flash('Booking berhasil diajukan. Menunggu approval HRD.', 'success')
            return redirect('/meeting-room/list')

        except BookingValidationError as e:
            flash(str(e), 'danger')
        except ValueError:
            flash('Format tanggal, jam, atau jumlah peserta tidak valid.', 'danger')

    return render_template('meeting_room/form.html', user=user, rooms=rooms, min_days=min_days)


@meeting_room_bp.route('/detail/<int:booking_id>')
def detail_booking(booking_id):
    if not require_login():
        return redirect('/login')

    user = get_current_user()
    booking = RoomBooking.query.get_or_404(booking_id)

    if not is_supervisor(user) and booking.user_id != user.id:
        flash('Kamu tidak punya akses ke booking ini.', 'danger')
        return redirect('/meeting-room/list')

    return render_template('meeting_room/detail.html', user=user, booking=booking)


@meeting_room_bp.route('/cancel/<int:booking_id>', methods=['POST'])
def cancel_booking(booking_id):
    if not require_login():
        return redirect('/login')

    user = get_current_user()
    booking = RoomBooking.query.get_or_404(booking_id)

    if booking.user_id != user.id:
        flash('Kamu hanya bisa membatalkan booking milik sendiri.', 'danger')
        return redirect('/meeting-room/list')

    if booking.status != RoomBooking.STATUS_PENDING:
        flash('Hanya booking Pending yang bisa dibatalkan.', 'danger')
        return redirect('/meeting-room/list')

    booking.status = RoomBooking.STATUS_CANCELLED
    db.session.commit()

    flash('Booking berhasil dibatalkan.', 'success')
    return redirect('/meeting-room/list')


@meeting_room_bp.route('/approval')
def approval_list():
    if not require_login():
        return redirect('/login')

    user = get_current_user()

    if not is_supervisor(user):
        flash(
            'Kamu tidak punya akses ke halaman approval.',
            'danger'
        )
        return redirect('/meeting-room/list')

    allowed_filters = (
        RoomBooking.VALID_STATUSES + ['All']
    )

    status_filter = request.args.get(
        'status',
        RoomBooking.STATUS_PENDING
    )

    if status_filter not in allowed_filters:
        status_filter = RoomBooking.STATUS_PENDING

    query = RoomBooking.query

    if status_filter != 'All':
        query = query.filter(
            RoomBooking.status == status_filter
        )

    if status_filter == RoomBooking.STATUS_PENDING:
        query = query.order_by(
            RoomBooking.created_at.asc()
        )
    else:
        query = query.order_by(
            RoomBooking.created_at.desc()
        )

    bookings = query.all()

    # Mengambil user sekaligus agar halaman lebih ringan
    user_ids = {
        booking.user_id
        for booking in bookings
        if booking.user_id
    }

    if user_ids:
        booking_users = User.query.filter(
            User.id.in_(user_ids)
        ).all()

        users_by_id = {
            booking_user.id: booking_user
            for booking_user in booking_users
        }
    else:
        users_by_id = {}

    status_counts = {
        RoomBooking.STATUS_PENDING:
            RoomBooking.query.filter_by(
                status=RoomBooking.STATUS_PENDING
            ).count(),

        RoomBooking.STATUS_APPROVED:
            RoomBooking.query.filter_by(
                status=RoomBooking.STATUS_APPROVED
            ).count(),

        RoomBooking.STATUS_REJECTED:
            RoomBooking.query.filter_by(
                status=RoomBooking.STATUS_REJECTED
            ).count(),

        RoomBooking.STATUS_CANCELLED:
            RoomBooking.query.filter_by(
                status=RoomBooking.STATUS_CANCELLED
            ).count(),
    }

    return render_template(
        'meeting_room/approval.html',
        user=user,
        bookings=bookings,
        users_by_id=users_by_id,
        current_filter=status_filter,
        status_counts=status_counts,
        total_count=RoomBooking.query.count(),
    )


@meeting_room_bp.route('/approve/<int:booking_id>', methods=['POST'])
def approve_booking(booking_id):
    if not require_login():
        return redirect('/login')

    user = get_current_user()
    if not is_hrd(user):
        flash('Hanya HRD yang bisa approve booking.', 'danger')
        return redirect('/meeting-room/approval')

    booking = RoomBooking.query.get_or_404(booking_id)

    if booking.user_id == user.id:
        flash('HRD tidak bisa approve booking milik sendiri.', 'danger')
        return redirect('/meeting-room/approval')

    if booking.status != RoomBooking.STATUS_PENDING:
        flash('Booking ini sudah diproses.', 'warning')
        return redirect('/meeting-room/approval')

    if has_schedule_conflict(
        booking.room_id,
        booking.meeting_date,
        booking.start_time,
        booking.end_time,
        exclude_booking_id=booking.id,
        statuses=[RoomBooking.STATUS_APPROVED],
    ):
        flash('Tidak bisa approve. Ruangan sudah terpakai di jam tersebut.', 'danger')
        return redirect('/meeting-room/approval')

    booking.status = RoomBooking.STATUS_APPROVED
    booking.approved_by = user.id
    booking.approved_at = datetime.utcnow()
    booking.reject_reason = None
    db.session.commit()

    flash('Booking berhasil di-approve.', 'success')
    return redirect('/meeting-room/approval')


@meeting_room_bp.route('/reject/<int:booking_id>', methods=['POST'])
def reject_booking(booking_id):
    if not require_login():
        return redirect('/login')

    user = get_current_user()
    if not is_hrd(user):
        flash('Hanya HRD yang bisa reject booking.', 'danger')
        return redirect('/meeting-room/approval')

    booking = RoomBooking.query.get_or_404(booking_id)

    if booking.user_id == user.id:
        flash('HRD tidak bisa reject booking milik sendiri.', 'danger')
        return redirect('/meeting-room/approval')

    if booking.status != RoomBooking.STATUS_PENDING:
        flash('Booking ini sudah diproses.', 'warning')
        return redirect('/meeting-room/approval')

    reason = request.form.get('reject_reason', '').strip()
    if not reason:
        flash('Alasan reject wajib diisi.', 'danger')
        return redirect('/meeting-room/approval')

    booking.status = RoomBooking.STATUS_REJECTED
    booking.approved_by = user.id
    booking.approved_at = datetime.utcnow()
    booking.reject_reason = reason
    db.session.commit()

    flash('Booking berhasil ditolak.', 'success')
    return redirect('/meeting-room/approval')


@meeting_room_bp.route('/rooms')
def rooms_list():
    if not require_login():
        return redirect('/login')

    user = get_current_user()
    if not is_supervisor(user):
        flash('Kamu tidak punya akses ke master ruangan.', 'danger')
        return redirect('/meeting-room/list')

    rooms = MeetingRoom.query.order_by(MeetingRoom.room_name.asc()).all()
    return render_template('meeting_room/rooms.html', user=user, rooms=rooms)


@meeting_room_bp.route('/rooms/create', methods=['GET', 'POST'])
def room_create():
    if not require_login():
        return redirect('/login')

    user = get_current_user()
    if not is_supervisor(user):
        flash('Kamu tidak punya akses ke master ruangan.', 'danger')
        return redirect('/meeting-room/list')

    if request.method == 'POST':
        room_name = request.form.get('room_name', '').strip()
        capacity = request.form.get('capacity', type=int)
        location = request.form.get('location', '').strip()
        facilities = request.form.get('facilities', '').strip()

        if not room_name or not capacity or capacity < 1 or not location:
            flash('Nama ruangan, kapasitas, dan lokasi wajib diisi dengan benar.', 'danger')
            return render_template('meeting_room/room_form.html', user=user, room=None)

        room = MeetingRoom(
            room_name=room_name,
            capacity=capacity,
            location=location,
            facilities=facilities,
            is_active=True,
        )
        db.session.add(room)
        db.session.commit()

        flash('Ruangan berhasil ditambahkan.', 'success')
        return redirect('/meeting-room/rooms')

    return render_template('meeting_room/room_form.html', user=user, room=None)


@meeting_room_bp.route('/rooms/edit/<int:room_id>', methods=['GET', 'POST'])
def room_edit(room_id):
    if not require_login():
        return redirect('/login')

    user = get_current_user()
    if not is_supervisor(user):
        flash('Kamu tidak punya akses ke master ruangan.', 'danger')
        return redirect('/meeting-room/list')

    room = MeetingRoom.query.get_or_404(room_id)

    if request.method == 'POST':
        room.room_name = request.form.get('room_name', '').strip()
        room.capacity = request.form.get('capacity', type=int)
        room.location = request.form.get('location', '').strip()
        room.facilities = request.form.get('facilities', '').strip()

        if not room.room_name or not room.capacity or room.capacity < 1 or not room.location:
            flash('Nama ruangan, kapasitas, dan lokasi wajib diisi dengan benar.', 'danger')
            return render_template('meeting_room/room_form.html', user=user, room=room)

        db.session.commit()
        flash('Ruangan berhasil diedit.', 'success')
        return redirect('/meeting-room/rooms')

    return render_template('meeting_room/room_form.html', user=user, room=room)


@meeting_room_bp.route('/rooms/toggle/<int:room_id>', methods=['POST'])
def room_toggle(room_id):
    if not require_login():
        return redirect('/login')

    user = get_current_user()
    if not is_supervisor(user):
        flash('Kamu tidak punya akses ke master ruangan.', 'danger')
        return redirect('/meeting-room/list')

    room = MeetingRoom.query.get_or_404(room_id)
    room.is_active = not room.is_active
    db.session.commit()

    flash('Status ruangan berhasil diperbarui.', 'success')
    return redirect('/meeting-room/rooms')


@meeting_room_bp.route('/schedule')
def schedule():
    if not require_login():
        return redirect('/login')

    user = get_current_user()

    current_date_str = request.args.get('date', '')
    room_id = request.args.get('room_id', type=int)
    status = request.args.get(
        'status',
        RoomBooking.STATUS_APPROVED
    )

    try:
        current_date = (
            parse_date(current_date_str)
            if current_date_str
            else date.today()
        )
    except ValueError:
        current_date = date.today()

    # Menentukan hari Senin dan Minggu pada minggu aktif
    week_start = current_date - timedelta(
        days=current_date.weekday()
    )
    week_end = week_start + timedelta(days=6)

    query = RoomBooking.query.filter(
        RoomBooking.meeting_date >= week_start,
        RoomBooking.meeting_date <= week_end
    )

    if room_id:
        query = query.filter(
            RoomBooking.room_id == room_id
        )

    if status and status != 'All':
        query = query.filter(
            RoomBooking.status == status
        )

    bookings = query.order_by(
        RoomBooking.meeting_date.asc(),
        RoomBooking.start_time.asc()
    ).all()

    rooms = MeetingRoom.get_active_rooms()

    # Mengelompokkan booking berdasarkan tanggal
    bookings_by_date = {
        week_start + timedelta(days=index): []
        for index in range(7)
    }

    for booking in bookings:
        if booking.meeting_date in bookings_by_date:
            bookings_by_date[booking.meeting_date].append(
                booking
            )

    day_names = [
        'Senin',
        'Selasa',
        'Rabu',
        'Kamis',
        'Jumat',
        'Sabtu',
        'Minggu'
    ]

    month_names = [
        'Januari',
        'Februari',
        'Maret',
        'April',
        'Mei',
        'Juni',
        'Juli',
        'Agustus',
        'September',
        'Oktober',
        'November',
        'Desember'
    ]

    week_days = []

    for index in range(7):
        day_date = week_start + timedelta(days=index)

        week_days.append({
            'date': day_date,
            'date_iso': day_date.isoformat(),
            'day_name': day_names[index],
            'date_number': day_date.day,
            'month_name': month_names[day_date.month - 1],
            'is_today': day_date == date.today(),
            'is_selected': day_date == current_date,
            'bookings': bookings_by_date[day_date]
        })

    if (
        week_start.month == week_end.month
        and week_start.year == week_end.year
    ):
        week_label = (
            f'{week_start.day}–{week_end.day} '
            f'{month_names[week_end.month - 1]} '
            f'{week_end.year}'
        )
    elif week_start.year == week_end.year:
        week_label = (
            f'{week_start.day} '
            f'{month_names[week_start.month - 1]} – '
            f'{week_end.day} '
            f'{month_names[week_end.month - 1]} '
            f'{week_end.year}'
        )
    else:
        week_label = (
            f'{week_start.day} '
            f'{month_names[week_start.month - 1]} '
            f'{week_start.year} – '
            f'{week_end.day} '
            f'{month_names[week_end.month - 1]} '
            f'{week_end.year}'
        )

    approved_count = sum(
        1 for booking in bookings
        if booking.status == RoomBooking.STATUS_APPROVED
    )

    pending_count = sum(
        1 for booking in bookings
        if booking.status == RoomBooking.STATUS_PENDING
    )

    return render_template(
        'meeting_room/schedule.html',
        user=user,
        bookings=bookings,
        rooms=rooms,
        current_date=current_date,
        room_id=room_id,
        status=status,
        week_days=week_days,
        week_label=week_label,
        previous_week=(week_start - timedelta(days=7)).isoformat(),
        next_week=(week_start + timedelta(days=7)).isoformat(),
        today_iso=date.today().isoformat(),
        approved_count=approved_count,
        pending_count=pending_count,
    )


@meeting_room_bp.route('/report')
def report():
    if not require_login():
        return redirect('/login')

    user = get_current_user()
    if not is_supervisor(user):
        flash('Kamu tidak punya akses ke laporan booking.', 'danger')
        return redirect('/meeting-room/list')

    start_date_str = request.args.get('start_date', '')
    end_date_str = request.args.get('end_date', '')
    status_filter = request.args.get('status', '')
    room_id = request.args.get('room_id', type=int)

    query = RoomBooking.query

    if start_date_str:
        try:
            query = query.filter(RoomBooking.meeting_date >= parse_date(start_date_str))
        except ValueError:
            pass

    if end_date_str:
        try:
            query = query.filter(RoomBooking.meeting_date <= parse_date(end_date_str))
        except ValueError:
            pass

    if status_filter:
        query = query.filter(RoomBooking.status == status_filter)

    if room_id:
        query = query.filter(RoomBooking.room_id == room_id)

    bookings = query.order_by(RoomBooking.meeting_date.desc(), RoomBooking.start_time.asc()).all()
    rooms = MeetingRoom.get_active_rooms()

    summary = {
        'total': len(bookings),
        'pending': sum(1 for b in bookings if b.status == RoomBooking.STATUS_PENDING),
        'approved': sum(1 for b in bookings if b.status == RoomBooking.STATUS_APPROVED),
        'rejected': sum(1 for b in bookings if b.status == RoomBooking.STATUS_REJECTED),
        'cancelled': sum(1 for b in bookings if b.status == RoomBooking.STATUS_CANCELLED),
    }

    return render_template(
        'meeting_room/report.html',
        user=user,
        bookings=bookings,
        rooms=rooms,
        statuses=RoomBooking.VALID_STATUSES,
        summary=summary,
        start_date=start_date_str,
        end_date=end_date_str,
        status_filter=status_filter,
        room_id_filter=room_id,
    )


@meeting_room_bp.route('/export_excel')
def export_excel():
    if not require_login():
        return redirect('/login')

    user = get_current_user()
    if not is_supervisor(user):
        flash('Kamu tidak punya akses export data booking.', 'danger')
        return redirect('/meeting-room/list')

    bookings = RoomBooking.query.order_by(RoomBooking.meeting_date.desc(), RoomBooking.start_time.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Booking Ruang Meeting'

    headers = [
        'No', 'Tanggal Pengajuan', 'Pemohon', 'Divisi', 'Ruangan', 'Lokasi',
        'Judul Meeting', 'Tanggal Meeting', 'Jam Mulai', 'Jam Selesai',
        'Peserta', 'Keperluan', 'Catatan', 'Status', 'Diproses Oleh',
        'Tanggal Proses', 'Alasan Reject'
    ]

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_index, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_index, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_index, booking in enumerate(bookings, 2):
        pemohon = get_user_by_id(booking.user_id)
        approver = get_user_by_id(booking.approved_by)
        row_values = [
            row_index - 1,
            booking.created_at.strftime('%Y-%m-%d %H:%M') if booking.created_at else '',
            display_name(pemohon),
            booking.division,
            booking.room.room_name if booking.room else '-',
            booking.room.location if booking.room else '-',
            booking.title,
            booking.meeting_date.strftime('%Y-%m-%d') if booking.meeting_date else '',
            booking.start_time.strftime('%H:%M') if booking.start_time else '',
            booking.end_time.strftime('%H:%M') if booking.end_time else '',
            booking.participant_count,
            booking.purpose,
            booking.notes,
            booking.status,
            display_name(approver),
            booking.approved_at.strftime('%Y-%m-%d %H:%M') if booking.approved_at else '',
            booking.reject_reason or '',
        ]

        for col_index, value in enumerate(row_values, 1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = thin_border

    for col_index in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_index).column_letter].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='data_booking_ruang_meeting.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
